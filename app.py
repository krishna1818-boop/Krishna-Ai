import os
from dotenv import load_dotenv
import pandas as pd
import streamlit as st
from openpyxl import Workbook, load_workbook
from twilio.rest import Client

twilio_client = Client(os.getenv("TWILIO_ACCOUNT_SID"), os.getenv("TWILIO_AUTH_TOKEN"))

st.set_page_config(page_title="Krishna AI", page_icon="🏠")
st.title("🏠 Krishna AI - Real Estate Assistant")
st.write("Welcome Krishna!")

name = st.text_input("👤 Customer Name")
phone = st.text_input("📞 Mobile Number")

budget = st.number_input(
    "💰 Budget (Lakhs)",
    min_value=10,
    max_value=500,
    value=35,
)

location = st.selectbox(
    "📍 Location",
    [
        "Dombivli",
        "Dombivli East",
        "Dombivli West",
        "Kalyan East",
        "Kharghar",
    ],
)

bhk = st.selectbox(
    "🏠 Select BHK",
    [
        "1 BHK",
        "2 BHK",
        "3 BHK",
    ],
)

requirement = st.text_area("📝 Customer Requirement")

if "project_result" not in st.session_state:
    st.session_state.project_result = pd.DataFrame()
    st.session_state.best_project = None
    st.session_state.sm = None

search_clicked = st.button("🔍 Find Best Project")
visit_clicked = st.button("📅 Book Site Visit")
save_clicked = st.button("💾 Save Lead")


def load_excel(path):

  from dotenv import load_dotenv
load_dotenv()

def send_whatsapp(to_number, message):
    st.write("SID:", os.getenv("TWILIO_ACCOUNT_SID"))
    st.write("WhatsApp:", os.getenv("TWILIO_WHATSAPP_NUMBER"))
    client = Client(
        os.getenv("TWILIO_ACCOUNT_SID"),
        os.getenv("TWILIO_AUTH_TOKEN")
    )

    client.messages.create(
        from_=os.getenv("TWILIO_WHATSAPP_NUMBER"),
        body=message,
        to=f"whatsapp:+91{to_number}"
    )

def load_excel(path):
    st.write(f"Loading: {path}")
    st.write(os.path.exists(path))

    if not os.path.exists(path):
        return None

    return pd.read_excel(path)

def ensure_lead_file(path):
    if not os.path.exists(path):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(["Name", "Mobile", "Budget", "Location", "BHK", "Requirement"])
        workbook.save(path)


if search_clicked:
    try:
        project_df = load_excel("Project.xlsx")
        
        sm_df = load_excel("SM_Details.xlsx")

        if project_df is None:
            st.error("❌ Could not find Project.xlsx in the app folder.")
        elif sm_df is None:
            st.error("❌ Could not find SM_Details.xlsx in the app folder.")
        else:
            filtered = project_df[
                (project_df["Location"] == location)
                & project_df["BHK"].str.contains(bhk.split()[0], case=False, na=False)
                & (project_df["Min Budget"] <= budget)
                & (project_df["Max Budget"] >= budget)
            ]

            if filtered.empty:
                st.error("❌ No Matching Project Found")
                st.session_state.project_result = pd.DataFrame()
                st.session_state.best_project = None
                st.session_state.sm = None
            else:
                filtered = filtered.copy()
                filtered["Score"] = 50 + (100 - abs(filtered["Min Budget"] - budget))
                filtered = filtered.sort_values(by="Score", ascending=False)
                st.session_state.project_result = filtered
                st.session_state.best_project = filtered.iloc[0]

                sm_match = sm_df[sm_df["Project"] == st.session_state.best_project["Project"]]
                st.session_state.sm = sm_match.iloc[0] if not sm_match.empty else None
    except Exception as e:
        st.error(f"Error loading data: {e}")
        st.session_state.project_result = pd.DataFrame()
        st.session_state.best_project = None
        st.session_state.sm = None

if not st.session_state.project_result.empty:
    st.success("✅ Matching Projects")
    st.dataframe(st.session_state.project_result)

    best = st.session_state.best_project
    st.success("🏆 Best Recommended Project")
    st.write(f"**Project:** {best['Project']}")
    st.write(f"**Location:** {best['Location']}")
    st.write(f"**Budget:** ₹{best['Min Budget']} - ₹{best['Max Budget']} Lakhs")
    st.write(f"**BHK:** {best['BHK']}")
    st.write(f"**Possession:** {best.get('Possession', 'N/A')}")
    st.write(f"**Amenities:** {best.get('Amenities', 'N/A')}")
    st.write(f"**Builder:** {best.get('Builder', 'N/A')}")

    if st.session_state.sm is not None:
        st.write(f"**Sales Manager:** {st.session_state.sm.get('SM Name', 'N/A')}")
        st.write(f"**SM Mobile:** {st.session_state.sm.get('SM Mobile', 'N/A')}")

    st.markdown("---")

    for _, row in st.session_state.project_result.iterrows():
        st.subheader(f"🏠 {row['Project']}")
        st.write(f"📍 Location: {row['Location']}")
        st.write(f"💰 Budget: ₹{row['Min Budget']} - ₹{row['Max Budget']} Lakhs")
        st.write(f"🛏️ BHK: {row['BHK']}")
        st.write(f"🏗️ Possession: {row.get('Possession', 'N/A')}")
        st.write(f"🚉 Station Distance: {row.get('Station Distance', 'N/A')}")
        st.write(f"✨ Amenities: {row.get('Amenities', 'N/A')}")
        st.write(f"🏢 Builder: {row.get('Builder', 'N/A')}")
        st.markdown("---")

if visit_clicked:
    if st.session_state.best_project is None:
        st.error("❌ Find a matching project before booking a site visit.")
    else:
        best = st.session_state.best_project

        st.success("✅ Site Visit Request Submitted")

        if not os.path.exists("Site_Visit.xlsx"):
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.append([
                "Customer",
                "Mobile",
                "Project",
                "Visit Date",
                "Visit Time",
                "Status",
            ])
        else:
            workbook = load_workbook("Site_Visit.xlsx")
            worksheet = workbook.active

        worksheet.append([name, phone, best["Project"], "", "", "Pending"])
        workbook.save("Site_Visit.xlsx")

        if st.session_state.sm is not None:
            message = f"""
New Site Visit Request

Customer: {name}
Mobile: {phone}
Project: {best['Project']}

Please contact the customer for the site visit.
"""

            send_whatsapp(
                str(st.session_state.sm["WhatsApp"]),
                message
            )

            st.success("📲 WhatsApp notification sent to Sales Manager.")

        st.write(f"Thank you, {name}!")
        st.write(f"🏠 Project: {best['Project']}")

        if st.session_state.sm is not None:
            st.write(f"📞 Relationship Manager: {st.session_state.sm.get('SM Name', 'N/A')}")
            st.write(f"📱 SM Mobile: {st.session_state.sm.get('SM Mobile', 'N/A')}")

        st.write("📞 Our team will contact you shortly to confirm your visit.")
        
if save_clicked:
    if st.session_state.best_project is None:
        st.error("❌ Find a project before saving a lead.")
    else:
        lead_file = "Leads.xlsx"
        ensure_lead_file(lead_file)

        workbook = load_workbook(lead_file)
        worksheet = workbook.active
        worksheet.append([name, phone, budget, location, bhk, requirement])
        workbook.save(lead_file)

        st.success("✅ Lead Saved Successfully!")
        st.subheader("📲 WhatsApp Message")
        message = f"""
Hello {name},

Thank you for your interest in Mumbai Homes.

📍 Location: {location}
💰 Budget: ₹{budget} Lakhs
🏠 Requirement: {requirement}

We have shortlisted suitable projects for you and will contact you shortly to schedule a site visit.
"""
        st.text_area("Copy this message and send on WhatsApp", message, height=200)

        st.subheader("🤖 Krishna AI Recommendation")
        if budget <= 50:
            st.success("🏠 Recommended: 1 BHK Projects")
        elif budget <= 70:
            st.success("🏠 Recommended: 2 BHK Projects")
        else:
            st.success("🏠 Recommended: Premium 2/3 BHK Projects")

        st.write(f"📍 Preferred Location: {location}")
        if "station" in requirement.lower():
            st.write("🚉 Priority: Projects near Railway Station")
        if "school" in requirement.lower():
            st.write("🏫 Priority: Projects near Schools")
        if "hospital" in requirement.lower():
            st.write("🏥 Priority: Projects near Hospital")
        st.info("📞 Next Step: Call the customer and schedule a site visit.")

st.subheader("🤖 AI Summary")
st.write(f"**Customer:** {name}")
st.write(f"**Mobile:** {phone}")
st.write(f"**Budget:** ₹{budget} Lakhs")
st.write(f"**Location:** {location}")
st.write(f"**Requirement:** {requirement}")

if st.session_state.best_project is not None:
    best = st.session_state.best_project
    recommendation = f"""
✅ Best Project: {best['Project']}

📍 Location: {best['Location']}
🏠 BHK: {best['BHK']}
💰 Budget: ₹{best['Min Budget']} - ₹{best['Max Budget']} Lakhs
🏗️ Possession: {best.get('Possession', 'N/A')}
🚉 Station Distance: {best.get('Station Distance', 'N/A')}
✨ Amenities: {best.get('Amenities', 'N/A')}
🏢 Builder: {best.get('Builder', 'N/A')}

This project is the best match based on your budget, location and BHK requirement.
"""
    st.info(recommendation)
else:
    st.info("Select budget, location and BHK, then click 'Find Best Project' to get a recommendation.")
st.markdown("---")
st.header("📋 Admin Dashboard")

st.markdown("## 📊 Dashboard Analytics")

total_leads = len(pd.read_excel("Leads.xlsx"))
total_visits = len(pd.read_excel("Site_Visit.xlsx"))

st.metric("👥 Total Leads", total_leads)
st.metric("🏠 Site Visits", total_visits)

visit_df = pd.read_excel("Site_Visit.xlsx")

confirmed = len(visit_df[visit_df["Status"] == "Confirmed"])
visited = len(visit_df[visit_df["Status"] == "Visited"])
cancelled = len(visit_df[visit_df["Status"] == "Cancelled"])

st.metric("✅ Confirmed Visits", confirmed)
st.metric("🏁 Completed Visits", visited)
st.metric("❌ Cancelled Visits", cancelled)

visit_df = pd.read_excel("Site_Visit.xlsx")

confirmed = len(visit_df[visit_df["Status"] == "Confirmed"])
visited = len(visit_df[visit_df["Status"] == "Visited"])
cancelled = len(visit_df[visit_df["Status"] == "Cancelled"])

st.metric("✅ Confirmed Visits", confirmed)
st.metric("🏁 Completed Visits", visited)
st.metric("❌ Cancelled Visits", cancelled)

try:
    leads_df = pd.read_excel("Leads.xlsx")
    st.subheader("👥 Customer Leads")
    st.dataframe(leads_df)
except:
    st.warning("No Leads Found")

try:
    visit_df = pd.read_excel("Site_Visit.xlsx")
    st.subheader("🏠 Site Visit Requests")
    st.dataframe(visit_df)
    st.subheader("🔄 Update Site Visit Status")

    if not visit_df.empty:
        row = st.selectbox(
            "Select Customer",
            visit_df.index,
            format_func=lambda x: f"{visit_df.loc[x, 'Customer']} - {visit_df.loc[x, 'Project']}"
        )

        new_status = st.selectbox(
            "New Status",
            ["Pending", "Confirmed", "Visited", "Cancelled"]
        )

        st.markdown("---")
        st.subheader("📞 Today's Follow-up Calls")

        if "Status" in visit_df.columns:
            pending = visit_df[visit_df["Status"] == "Pending"]
        else:
            pending = pd.DataFrame()

        if pending.empty:
            st.success("✅ No Pending Follow-ups")
        else:
            st.dataframe(pending)

        if st.button("✅ Update Status"):
            visit_df.loc[row, "Status"] = new_status
            visit_df.to_excel("Site_Visit.xlsx", index=False)
            st.success("Status Updated Successfully!")
            st.subheader("💬 Generate Follow-up Message")
    else:
        st.warning("No Site Visit Requests Found")
except Exception:
    st.warning("No Site Visit Requests Found")

customer = st.text_input("Customer Name", key="followup_customer")
status = st.selectbox(
    "Follow-up Type",
    [
        "New Inquiry",
        "Site Visit Pending",
        "Site Visit Done",
        "Interested",
        "Not Interested"
    ]
)

if st.button("Generate Message"):
    if status == "New Inquiry":
        msg = f"Hi {customer}, thank you for your property inquiry. Aapki requirement ke hisaab se best options share karta hu."
    elif status == "Site Visit Pending":
        msg = f"Hi {customer}, aapke liye site visit schedule karna tha. Aap kis time convenient rahoge?"
    elif status == "Site Visit Done":
        msg = f"Hi {customer}, site visit ke liye thank you. Aapko property kaisi lagi? Aapke feedback ka wait rahega."
    elif status == "Interested":
        msg = f"Hi {customer}, aapke selected property ke next steps aur offers discuss kar sakte hain."
    else:
        msg = f"Hi {customer}, future me agar koi suitable property requirement ho to zarur bataye."

    st.success(msg)